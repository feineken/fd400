"""
FD400 — Generador de Nominas Banco de Chile
Interfaz web para generar archivos FD400 desde un Excel.
"""

import io
import streamlit as st
import openpyxl
from datetime import datetime


# ── password gate ─────────────────────────────────────────────────────────────

def check_password():
    """Simple password gate using Streamlit secrets. Skip if no secret set."""
    if "password" not in st.secrets:
        return True
    if st.session_state.get("authenticated"):
        return True

    st.title("FD400 — Banco de Chile")
    pwd = st.text_input("Contraseña", type="password")
    if st.button("Ingresar"):
        if pwd == st.secrets["password"]:
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("Contraseña incorrecta")
    return False


# ── helpers ───────────────────────────────────────────────────────────────────

def pad_num(v, n):
    return str(v).strip().zfill(n)

def pad_char(v, n):
    s = str(v).strip() if v is not None else ""
    return s.ljust(n)[:n]

def fmt_monto(pesos):
    cents = int(round(float(pesos) * 100))
    return str(cents).zfill(13)


# ── registros remuneraciones ──────────────────────────────────────────────────

def rem_reg1(cfg, total):
    r  = "01"
    r += pad_num(cfg["rut"], 9)
    r += pad_char(cfg["dv"], 1)
    r += pad_num(cfg["convenio"], 3)
    r += pad_num(cfg["num_nomina"], 5)
    r += pad_char(cfg["nombre_nomina"], 25)
    r += "01"
    r += cfg["fecha_pago"]
    r += fmt_monto(total)
    r += " " * 3
    r += "N"
    r += " " * 322
    r += "01"
    r += "0101"
    assert len(r) == 400, f"Reg1 largo incorrecto: {len(r)}"
    return r

def rem_reg2(cfg, row, num_msg):
    r  = "02"
    r += pad_num(cfg["rut"], 9)
    r += pad_char(cfg["dv"], 1)
    r += pad_num(cfg["convenio"], 3)
    r += "  "
    r += pad_num(cfg["num_nomina"], 5)
    r += pad_num(row[5], 2)
    r += pad_num(row[0], 9)
    r += pad_char(row[1], 1)
    r += pad_char(row[2], 60)
    r += "0"
    r += " " * 35
    r += " " * 15
    r += " " * 15
    r += " " * 7
    r += "  "
    r += pad_num(row[3], 3)
    r += pad_char(row[4], 22)
    r += "000"
    r += fmt_monto(row[6])
    r += pad_char(row[7], 119)
    r += pad_num(num_msg, 4)
    r += "N"
    r += "   "
    r += "0000000000"
    r += "+"
    r += "000000"
    r += "N"
    r += " " * 45
    assert len(r) == 400, f"Reg2 fila {num_msg} largo incorrecto: {len(r)}"
    return r

def rem_reg3(cfg, row, num_msg):
    glosa = row[9] if len(row) > 9 and row[9] is not None else ""
    r  = "03"
    r += pad_num(cfg["rut"], 9)
    r += pad_char(cfg["dv"], 1)
    r += pad_num(cfg["convenio"], 3)
    r += "  "
    r += pad_num(cfg["num_nomina"], 5)
    r += pad_num(num_msg, 4)
    r += "EMA"
    r += pad_char(row[8], 80)
    r += " " * 16
    r += pad_char(glosa, 250)
    r += "  "
    r += "000"
    r += " " * 20
    assert len(r) == 400, f"Reg3 fila {num_msg} largo incorrecto: {len(r)}"
    return r


# ── registros proveedores ─────────────────────────────────────────────────────

def prov_reg1(cfg, total):
    r  = "01"
    r += pad_num(cfg["rut"], 9)
    r += pad_char(cfg["dv"], 1)
    r += pad_num(cfg["convenio"], 3)
    r += pad_num(cfg["num_nomina"], 5)
    r += pad_char(cfg["nombre_nomina"], 25)
    r += "01"
    r += cfg["fecha_pago"]
    r += fmt_monto(total)
    r += " " * 3
    r += "N"
    r += " " * 322
    r += "01"
    r += "0202"
    assert len(r) == 400, f"Reg1 largo incorrecto: {len(r)}"
    return r

def prov_reg2(cfg, row, num_msg):
    r  = "02"
    r += pad_num(cfg["rut"], 9)
    r += pad_char(cfg["dv"], 1)
    r += pad_num(cfg["convenio"], 3)
    r += "  "
    r += pad_num(cfg["num_nomina"], 5)
    r += pad_num(row[5], 2)
    r += pad_num(row[0], 9)
    r += pad_char(row[1], 1)
    r += pad_char(row[2], 60)
    r += "0"
    r += " " * 35
    r += " " * 15
    r += " " * 15
    r += " " * 7
    r += "  "
    r += pad_num(row[3], 3)
    r += pad_char(row[4], 22)
    r += "   "
    r += fmt_monto(row[6])
    r += pad_char(row[7], 119)
    r += "0000"
    r += "N"
    r += "   "
    r += "          "
    r += " "
    r += pad_num(num_msg, 6)
    r += "N"
    r += " " * 45
    assert len(r) == 400, f"Reg2 proveedor {num_msg} largo incorrecto: {len(r)}"
    return r

def prov_reg3(cfg, row, num_msg, num_reg4):
    r  = "03"
    r += pad_num(cfg["rut"], 9)
    r += pad_char(cfg["dv"], 1)
    r += pad_num(cfg["convenio"], 3)
    r += "  "
    r += pad_num(cfg["num_nomina"], 5)
    r += pad_num(num_msg, 4)
    r += "EMA"
    r += pad_char(row[8], 80)
    r += " " * 16
    r += pad_char(row[9], 250)
    r += "  "
    r += pad_num(num_reg4, 3)
    r += " " * 20
    assert len(r) == 400, f"Reg3 proveedor {num_msg} largo incorrecto: {len(r)}"
    return r

def prov_reg4(cfg, row, num_msg, correlativo):
    cartola = ""
    for i in range(4):
        g_idx = 10 + i * 2
        m_idx = 11 + i * 2
        glosa = row[g_idx] if len(row) > g_idx and row[g_idx] is not None else ""
        monto = row[m_idx] if len(row) > m_idx and row[m_idx] is not None else 0
        cartola += pad_char(glosa, 67)
        cartola += fmt_monto(monto)
    r  = "04"
    r += pad_num(cfg["rut"], 9)
    r += pad_char(cfg["dv"], 1)
    r += pad_num(cfg["convenio"], 3)
    r += "  "
    r += pad_num(cfg["num_nomina"], 5)
    r += pad_num(num_msg, 4)
    r += cartola
    r += " " * 51
    r += pad_num(correlativo, 3)
    assert len(r) == 400, f"Reg4 proveedor {num_msg} largo incorrecto: {len(r)}"
    return r


# ── generacion en memoria ─────────────────────────────────────────────────────

def generar_remuneraciones(cfg, file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["Remuneraciones"]
    rows, total = [], 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rows.append(row)
        total += float(row[6])
    if not rows:
        raise ValueError("No se encontraron filas con datos en la hoja 'Remuneraciones'.")
    lines = [rem_reg1(cfg, total)]
    for i, row in enumerate(rows, 1):
        lines.append(rem_reg2(cfg, row, i))
        if len(row) > 8 and row[8] is not None and str(row[8]).strip():
            lines.append(rem_reg3(cfg, row, i))
    content = "\r\n".join(lines) + "\r\n"
    return content.encode("utf-8"), len(rows), total


def generar_proveedores(cfg, file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["Proveedores"]
    rows, total = [], 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rows.append(row)
        total += float(row[6])
    if not rows:
        raise ValueError("No se encontraron filas con datos en la hoja 'Proveedores'.")
    lines = [prov_reg1(cfg, total)]
    for i, row in enumerate(rows, 1):
        hay_email    = len(row) > 8 and row[8] is not None and str(row[8]).strip()
        hay_facturas = len(row) > 10 and row[10] is not None and str(row[10]).strip()
        num_reg4 = 1 if hay_facturas else 0
        lines.append(prov_reg2(cfg, row, i))
        if hay_email:
            lines.append(prov_reg3(cfg, row, i, num_reg4))
        if hay_facturas:
            lines.append(prov_reg4(cfg, row, i, 1))
    content = "\r\n".join(lines) + "\r\n"
    return content.encode("utf-8"), len(rows), total


# ── form de empresa ───────────────────────────────────────────────────────────

def empresa_form(prefix):
    """Renders company fields and returns a cfg dict. Returns None if incomplete."""
    col1, col2 = st.columns(2)
    with col1:
        rut = st.text_input("RUT Empresa (sin DV, sin puntos)", key=f"{prefix}_rut")
        dv  = st.text_input("DV Empresa", max_chars=1, key=f"{prefix}_dv")
        convenio = st.text_input("Código Convenio (3 dígitos)", max_chars=3, key=f"{prefix}_conv")
    with col2:
        num_nomina    = st.text_input("N° Nómina (max 5 dígitos)", max_chars=5, key=f"{prefix}_num")
        nombre_nomina = st.text_input("Nombre Nómina (max 25 caracteres)", max_chars=25, key=f"{prefix}_nombre")
        fecha_pago    = st.text_input("Fecha de Pago (AAAAMMDD)", max_chars=8,
                                      value=datetime.now().strftime("%Y%m%d"), key=f"{prefix}_fecha")
    return {
        "rut": rut,
        "dv": dv,
        "convenio": convenio,
        "num_nomina": num_nomina,
        "nombre_nomina": nombre_nomina,
        "fecha_pago": fecha_pago,
    }


def validar_cfg(cfg):
    for k, v in cfg.items():
        if not str(v).strip():
            return f"El campo '{k}' no puede estar vacío."
    if len(cfg["fecha_pago"]) != 8 or not cfg["fecha_pago"].isdigit():
        return "La fecha debe tener exactamente 8 dígitos (AAAAMMDD)."
    return None


# ── UI principal ──────────────────────────────────────────────────────────────

def tab_remuneraciones():
    st.subheader("Datos de la empresa")
    cfg = empresa_form("rem")

    st.subheader("Archivo Excel")
    uploaded = st.file_uploader(
        "Sube el Excel de remuneraciones (hoja: Remuneraciones)",
        type=["xlsx"],
        key="rem_file",
    )

    if st.button("Generar archivo FD400", key="rem_gen"):
        err = validar_cfg(cfg)
        if err:
            st.error(err)
        elif uploaded is None:
            st.error("Debes subir el archivo Excel antes de continuar.")
        else:
            try:
                data, n, total = generar_remuneraciones(cfg, uploaded.read())
                fecha = datetime.now().strftime("%Y%m%d")
                st.success(f"Generado correctamente — {n} beneficiarios — Total: ${total:,.0f}")
                st.download_button(
                    label="Descargar remuneraciones.txt",
                    data=data,
                    file_name=f"remuneraciones_{fecha}.txt",
                    mime="text/plain",
                )
            except Exception as e:
                st.error(f"Error: {e}")


def tab_proveedores():
    st.subheader("Datos de la empresa")
    cfg = empresa_form("prov")

    st.subheader("Archivo Excel")
    uploaded = st.file_uploader(
        "Sube el Excel de proveedores (hoja: Proveedores)",
        type=["xlsx"],
        key="prov_file",
    )

    if st.button("Generar archivo FD400", key="prov_gen"):
        err = validar_cfg(cfg)
        if err:
            st.error(err)
        elif uploaded is None:
            st.error("Debes subir el archivo Excel antes de continuar.")
        else:
            try:
                data, n, total = generar_proveedores(cfg, uploaded.read())
                fecha = datetime.now().strftime("%Y%m%d")
                st.success(f"Generado correctamente — {n} proveedores — Total: ${total:,.0f}")
                st.download_button(
                    label="Descargar proveedores.txt",
                    data=data,
                    file_name=f"proveedores_{fecha}.txt",
                    mime="text/plain",
                )
            except Exception as e:
                st.error(f"Error: {e}")


def main():
    st.set_page_config(page_title="FD400 — Banco de Chile", page_icon="🏦", layout="centered")

    if not check_password():
        st.stop()

    st.title("Generador de Nóminas FD400")
    st.caption("Banco de Chile — Pagos Masivos")

    tab1, tab2 = st.tabs(["Remuneraciones", "Proveedores"])
    with tab1:
        tab_remuneraciones()
    with tab2:
        tab_proveedores()


if __name__ == "__main__":
    main()
