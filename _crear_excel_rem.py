import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment

wb = Workbook()

# ── Hoja Listas ─────────────────────────────────────────────────────────────
ls = wb.create_sheet("Listas")

bancos = [
    ("001", "Banco de Chile"),
    ("009", "Banco Internacional"),
    ("012", "Banco Estado"),
    ("014", "Scotiabank"),
    ("016", "BCI"),
    ("028", "BICE"),
    ("031", "HSBC"),
    ("037", "Santander"),
    ("039", "Itau / Corpbanca"),
    ("041", "JP Morgan"),
    ("049", "Security"),
    ("051", "Falabella"),
    ("053", "Banco Ripley"),
    ("055", "Consorcio"),
    ("672", "Coopeuch"),
    ("729", "Prepago Los Heroes"),
    ("730", "Tenpo Prepago"),
]

medios = [
    ("01", "Cta Cte Banco de Chile"),
    ("06", "Cta Ahorro Banco de Chile"),
    ("07", "Cta Cte otros bancos"),
    ("08", "Cta Ahorro otros bancos"),
    ("11", "Bancuenta Credichile"),
    ("12", "Pago efectivo Servipag"),
]

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF")

ls["A1"] = "Cod. Banco"
ls["B1"] = "Nombre Banco"
ls["D1"] = "Cod. Medio Pago"
ls["E1"] = "Descripcion"
for cell in [ls["A1"], ls["B1"], ls["D1"], ls["E1"]]:
    cell.fill = header_fill
    cell.font = header_font

for i, (cod, desc) in enumerate(bancos, 2):
    ls.cell(row=i, column=1, value=cod)
    ls.cell(row=i, column=2, value=desc)

for i, (cod, desc) in enumerate(medios, 2):
    ls.cell(row=i, column=4, value=cod)
    ls.cell(row=i, column=5, value=desc)

ls.column_dimensions["A"].width = 12
ls.column_dimensions["B"].width = 28
ls.column_dimensions["D"].width = 16
ls.column_dimensions["E"].width = 30

n_bancos = len(bancos) + 1
n_medios = len(medios) + 1

# ── Hoja Remuneraciones ─────────────────────────────────────────────────────
ws = wb.active
ws.title = "Remuneraciones"

headers = ["RUT", "DV", "Nombre", "Banco", "Cuenta", "MedioPago", "Monto", "Descripcion"]
ws.append(headers)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

anchos = [13, 4, 42, 8, 20, 12, 12, 30]
for i, w in enumerate(anchos, 1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

MAX_ROWS = 1000

dv_banco = DataValidation(
    type="list",
    formula1=f"Listas!$A$2:$A${n_bancos}",
    allow_blank=True,
    showErrorMessage=True,
    error="Seleccione un banco de la lista",
    errorTitle="Banco invalido",
)
dv_banco.sqref = f"D2:D{MAX_ROWS}"
ws.add_data_validation(dv_banco)

dv_medio = DataValidation(
    type="list",
    formula1=f"Listas!$D$2:$D${n_medios}",
    allow_blank=True,
    showErrorMessage=True,
    error="Seleccione un medio de pago de la lista",
    errorTitle="Medio invalido",
)
dv_medio.sqref = f"F2:F{MAX_ROWS}"
ws.add_data_validation(dv_medio)

ws.freeze_panes = "A2"

wb.save("remuneraciones_template.xlsx")
print("OK: remuneraciones_template.xlsx")
print("   Hoja: Remuneraciones | Columnas A-H | Dropdowns en D y F")
