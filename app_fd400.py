# app_fd400.py
# Interfaz grafica para generar archivos FD400 (Banco de Chile)
# Empaquetable como .exe con: pyinstaller --onefile --windowed app_fd400.py

import json
import os
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
import openpyxl


def get_base_path():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


CONFIG_FILE = os.path.join(get_base_path(), "fd400_config.json")


# ── helpers ───────────────────────────────────────────────────────────────────

def pad_num(v, n):
    return str(v).strip().zfill(n)

def pad_char(v, n):
    s = str(v).strip() if v is not None else ""
    return s.ljust(n)[:n]

def fmt_monto(pesos):
    cents = int(round(float(pesos) * 100))
    return str(cents).zfill(13)


# ── registros remuneraciones ──────────────────────────────────────────────────

def rem_reg1(cfg, total):
    r  = "01"
    r += pad_num(cfg["rut"], 9)
    r += pad_char(cfg["dv"], 1)
    r += pad_num(cfg["convenio"], 3)
    r += pad_num(cfg["num_nomina"], 5)
    r += pad_char(cfg["nombre_nomina"], 25)
    r += "01"
    r += cfg["fecha_pago"]
    r += fmt_monto(total)
    r += " " * 3
    r += "N"
    r += " " * 322
    r += "01"
    r += "0101"
    assert len(r) == 400, f"Reg1 largo incorrecto: {len(r)}"
    return r

def rem_reg2(cfg, row, num_msg):
    r  = "02"
    r += pad_num(cfg["rut"], 9)
    r += pad_char(cfg["dv"], 1)
    r += pad_num(cfg["convenio"], 3)
    r += "  "
    r += pad_num(cfg["num_nomina"], 5)
    r += pad_num(row[5], 2)
    r += pad_num(row[0], 9)
    r += pad_char(row[1], 1)
    r += pad_char(row[2], 60)
    r += "0"
    r += " " * 35
    r += " " * 15
    r += " " * 15
    r += " " * 7
    r += "  "
    r += pad_num(row[3], 3)
    r += pad_char(row[4], 22)
    r += "000"
    r += fmt_monto(row[6])
    r += pad_char(row[7], 119)
    r += pad_num(num_msg, 4)
    r += "N"
    r += "   "
    r += "0000000000"
    r += "+"
    r += "000000"
    r += "N"
    r += " " * 45
    assert len(r) == 400, f"Reg2 fila {num_msg} largo incorrecto: {len(r)}"
    return r

def rem_reg3(cfg, row, num_msg):
    glosa = row[9] if len(row) > 9 and row[9] is not None else ""
    r  = "03"
    r += pad_num(cfg["rut"], 9)
    r += pad_char(cfg["dv"], 1)
    r += pad_num(cfg["convenio"], 3)
    r += "  "
    r += pad_num(cfg["num_nomina"], 5)
    r += pad_num(num_msg, 4)
    r += "EMA"
    r += pad_char(row[8], 80)
    r += " " * 16
    r += pad_char(glosa, 250)
    r += "  "
    r += "000"
    r += " " * 20
    assert len(r) == 400, f"Reg3 fila {num_msg} largo incorrecto: {len(r)}"
    return r


# ── registros proveedores ─────────────────────────────────────────────────────

def prov_reg1(cfg, total):
    r  = "01"
    r += pad_num(cfg["rut"], 9)
    r += pad_char(cfg["dv"], 1)
    r += pad_num(cfg["convenio"], 3)
    r += pad_num(cfg["num_nomina"], 5)
    r += pad_char(cfg["nombre_nomina"], 25)
    r += "01"
    r += cfg["fecha_pago"]
    r += fmt_monto(total)
    r += " " * 3
    r += "N"
    r += " " * 322
    r += "01"
    r += "0202"
    assert len(r) == 400, f"Reg1 largo incorrecto: {len(r)}"
    return r

def prov_reg2(cfg, row, num_msg):
    r  = "02"
    r += pad_num(cfg["rut"], 9)
    r += pad_char(cfg["dv"], 1)
    r += pad_num(cfg["convenio"], 3)
    r += "  "
    r += pad_num(cfg["num_nomina"], 5)
    r += pad_num(row[5], 2)
    r += pad_num(row[0], 9)
    r += pad_char(row[1], 1)
    r += pad_char(row[2], 60)
    r += "0"
    r += " " * 35
    r += " " * 15
    r += " " * 15
    r += " " * 7
    r += "  "
    r += pad_num(row[3], 3)
    r += pad_char(row[4], 22)
    r += "   "
    r += fmt_monto(row[6])
    r += pad_char(row[7], 119)
    r += "0000"
    r += "N"
    r += "   "
    r += "          "
    r += " "
    r += pad_num(num_msg, 6)
    r += "N"
    r += " " * 45
    assert len(r) == 400, f"Reg2 proveedor {num_msg} largo incorrecto: {len(r)}"
    return r

def prov_reg3(cfg, row, num_msg, num_reg4):
    r  = "03"
    r += pad_num(cfg["rut"], 9)
    r += pad_char(cfg["dv"], 1)
    r += pad_num(cfg["convenio"], 3)
    r += "  "
    r += pad_num(cfg["num_nomina"], 5)
    r += pad_num(num_msg, 4)
    r += "EMA"
    r += pad_char(row[8], 80)
    r += " " * 16
    r += pad_char(row[9], 250)
    r += "  "
    r += pad_num(num_reg4, 3)
    r += " " * 20
    assert len(r) == 400, f"Reg3 proveedor {num_msg} largo incorrecto: {len(r)}"
    return r

def prov_reg4(cfg, row, num_msg, correlativo):
    cartola = ""
    for i in range(4):
        g_idx = 10 + i * 2
        m_idx = 11 + i * 2
        glosa = row[g_idx] if len(row) > g_idx and row[g_idx] is not None else ""
        monto = row[m_idx] if len(row) > m_idx and row[m_idx] is not None else 0
        cartola += pad_char(glosa, 67)
        cartola += fmt_monto(monto)
    r  = "04"
    r += pad_num(cfg["rut"], 9)
    r += pad_char(cfg["dv"], 1)
    r += pad_num(cfg["convenio"], 3)
    r += "  "
    r += pad_num(cfg["num_nomina"], 5)
    r += pad_num(num_msg, 4)
    r += cartola
    r += " " * 51
    r += pad_num(correlativo, 3)
    assert len(r) == 400, f"Reg4 proveedor {num_msg} largo incorrecto: {len(r)}"
    return r


# ── generacion ────────────────────────────────────────────────────────────────

def generar_remuneraciones(cfg, archivo_excel):
    wb = openpyxl.load_workbook(archivo_excel, data_only=True)
    ws = wb["Remuneraciones"]
    rows, total = [], 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rows.append(row)
        total += float(row[6])
    if not rows:
        raise ValueError("No se encontraron filas con datos en la hoja 'Remuneraciones'.")
    lines = [rem_reg1(cfg, total)]
    for i, row in enumerate(rows, 1):
        lines.append(rem_reg2(cfg, row, i))
        if len(row) > 8 and row[8] is not None and str(row[8]).strip():
            lines.append(rem_reg3(cfg, row, i))
    fecha = datetime.now().strftime("%Y%m%d")
    salida = os.path.join(os.path.dirname(archivo_excel), f"remuneraciones_{fecha}.txt")
    with open(salida, "w", encoding="utf-8", newline="") as f:
        f.write("\r\n".join(lines) + "\r\n")
    return salida, len(rows), total


def generar_proveedores(cfg, archivo_excel):
    wb = openpyxl.load_workbook(archivo_excel, data_only=True)
    ws = wb["Proveedores"]
    rows, total = [], 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rows.append(row)
        total += float(row[6])
    if not rows:
        raise ValueError("No se encontraron filas con datos en la hoja 'Proveedores'.")
    lines = [prov_reg1(cfg, total)]
    for i, row in enumerate(rows, 1):
        hay_email    = len(row) > 8 and row[8] is not None and str(row[8]).strip()
        hay_facturas = len(row) > 10 and row[10] is not None and str(row[10]).strip()
        num_reg4 = 1 if hay_facturas else 0
        lines.append(prov_reg2(cfg, row, i))
        if hay_email:
            lines.append(prov_reg3(cfg, row, i, num_reg4))
        if hay_facturas:
            lines.append(prov_reg4(cfg, row, i, 1))
    fecha = datetime.now().strftime("%Y%m%d")
    salida = os.path.join(os.path.dirname(archivo_excel), f"proveedores_{fecha}.txt")
    with open(salida, "w", encoding="utf-8", newline="") as f:
        f.write("\r\n".join(lines) + "\r\n")
    return salida, len(rows), total


# ── config persistente ────────────────────────────────────────────────────────

def load_config():
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def save_config(data):
    try:
        existing = load_config()
        existing.update(data)
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(existing, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# ── GUI ───────────────────────────────────────────────────────────────────────

FIELDS = [
    ("RUT Empresa  (sin DV, sin puntos)",    "rut",           ""),
    ("DV Empresa",                           "dv",            ""),
    ("Codigo Convenio",                      "convenio",      ""),
    ("N° Nomina  (max 5 digitos)",      "num_nomina",    ""),
    ("Nombre Nomina  (max 25 caracteres)",   "nombre_nomina", ""),
    ("Fecha de Pago  (AAAAMMDD)",            "fecha_pago",    ""),
]


class Tab(ttk.Frame):
    def __init__(self, parent, prefix, fn_generar):
        super().__init__(parent, padding=20)
        self.prefix = prefix
        self.fn_generar = fn_generar
        cfg = load_config()

        self.vars = {}
        for i, (label, key, default) in enumerate(FIELDS):
            ttk.Label(self, text=label, anchor="w").grid(row=i, column=0, sticky="w", pady=5, padx=(0, 12))
            stored = cfg.get(f"{prefix}_{key}", default)
            if key == "fecha_pago" and not stored:
                stored = datetime.now().strftime("%Y%m%d")
            var = tk.StringVar(value=stored)
            ttk.Entry(self, textvariable=var, width=30).grid(row=i, column=1, sticky="ew", pady=5)
            self.vars[key] = var

        sep_row = len(FIELDS)
        ttk.Separator(self, orient="horizontal").grid(row=sep_row, column=0, columnspan=2, sticky="ew", pady=12)

        ttk.Label(self, text="Archivo Excel", anchor="w").grid(row=sep_row+1, column=0, sticky="w", pady=5, padx=(0, 12))
        self.archivo_var = tk.StringVar()
        ttk.Entry(self, textvariable=self.archivo_var, width=30, state="readonly").grid(row=sep_row+1, column=1, sticky="ew", pady=5)
        ttk.Button(self, text="Buscar archivo...", command=self._browse).grid(row=sep_row+2, column=1, sticky="w", pady=(0, 12))

        btn = ttk.Button(self, text="Generar archivo FD400", command=self._generar)
        btn.grid(row=sep_row+3, column=0, columnspan=2, sticky="ew", pady=(4, 8), ipady=6)

        self.result_var = tk.StringVar()
        self.result_label = ttk.Label(self, textvariable=self.result_var, wraplength=420, justify="left")
        self.result_label.grid(row=sep_row+4, column=0, columnspan=2, sticky="w", pady=(4, 0))

        self.columnconfigure(1, weight=1)

    def _browse(self):
        path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos los archivos", "*.*")],
        )
        if path:
            self.archivo_var.set(path)

    def _generar(self):
        self.result_var.set("")
        archivo = self.archivo_var.get().strip()
        if not archivo:
            messagebox.showwarning("Falta el archivo", "Por favor selecciona el archivo Excel antes de continuar.")
            return

        cfg = {k: v.get().strip() for k, v in self.vars.items()}
        for label, key, _ in FIELDS:
            if not cfg[key]:
                messagebox.showwarning("Campo vacio", f"El campo '{label.split('  ')[0]}' no puede estar vacio.")
                return
        if len(cfg["fecha_pago"]) != 8 or not cfg["fecha_pago"].isdigit():
            messagebox.showwarning("Fecha invalida", "La fecha debe tener exactamente 8 digitos.\nEjemplo: 202XXXXX")
            return

        try:
            salida, n, total = self.fn_generar(cfg, archivo)
            save_config({f"{self.prefix}_{k}": v for k, v in cfg.items()})
            self.result_label.configure(foreground="#1a7a1a")
            self.result_var.set(
                f"Archivo generado correctamente:\n{salida}\n\n"
                f"Registros: {n}     Monto total: ${total:,.0f}"
            )
        except Exception as e:
            self.result_label.configure(foreground="#c0392b")
            self.result_var.set(f"Error: {e}")


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("FD400 — Banco de Chile")
        self.resizable(False, False)

        ttk.Label(self, text="Generador de Nominas FD400", font=("Segoe UI", 13, "bold")).pack(pady=(16, 4))
        ttk.Label(self, text="Banco de Chile — Pagos Masivos", foreground="#555").pack(pady=(0, 10))

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        tab_rem  = Tab(nb, "rem",  generar_remuneraciones)
        tab_prov = Tab(nb, "prov", generar_proveedores)
        nb.add(tab_rem,  text="   Remuneraciones   ")
        nb.add(tab_prov, text="   Proveedores   ")


if __name__ == "__main__":
    app = App()
    app.mainloop()
